import streamlit as st
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from fpdf import FPDF
import pandas as pd
import tempfile

st.title("Excel Cell Updater and PDF Plot Generator")

# Upload Excel file
uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx"])

if uploaded_file:
    # Load the workbook and select a sheet
    workbook = load_workbook(uploaded_file, data_only=True)
    sheet_names = workbook.sheetnames
    sheet_name = st.selectbox("Select Sheet", sheet_names)
    sheet = workbook[sheet_name]

    # Display the current data in the sheet
    st.write("Current Sheet Data:")
    data = sheet.values
    df = pd.DataFrame(data)
    st.write(df)

    # Input for the cell to update
    row = st.number_input("Row to Update", min_value=1, step=1)
    column = st.text_input("Column to Update (e.g., A, B, C)")

    # Input for multiple values
    values = st.text_area("Enter Values (one per line)").splitlines()

    if st.button("Generate PDFs"):
        if column and row > 0 and values:
            pdf_paths = []

            for value in values:
                # Update cell with the value
                cell = f"{column}{row}"
                sheet[cell] = value

                # Generate plot
                fig, ax = plt.subplots()
                ax.plot([1, 2, 3], [int(value), int(value)*2, int(value)*3], marker="o")
                ax.set_title(f"Plot for Value: {value}")
                ax.set_xlabel("X-axis")
                ax.set_ylabel("Y-axis")

                # Save each plot to a temporary PDF file
                with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_pdf:
                    pdf_path = tmp_pdf.name
                    fig.savefig(pdf_path, format="pdf")
                    pdf_paths.append(pdf_path)

                plt.close(fig)

            # Download each PDF
            for pdf_path in pdf_paths:
                with open(pdf_path, "rb") as file:
                    st.download_button(
                        label=f"Download PDF for Value {pdf_path.split('/')[-1].split('.')[0]}",
                        data=file,
                        file_name=f"Plot_{pdf_path.split('/')[-1].split('.')[0]}.pdf",
                        mime="application/pdf"
                    )
        else:
            st.warning("Please provide a valid column, row, and values.")
